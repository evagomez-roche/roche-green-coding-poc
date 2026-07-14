import csv
import os
import re
import time
import argparse
import platform
import multiprocessing
import openpyxl
import requests
from PyPDF2 import PdfReader
from codecarbon import EmissionsTracker
from datetime import datetime

# =====================================================================
# ⚙️ DYNAMIC CONFIGURATION & AUTOMATION
# =====================================================================
parser = argparse.ArgumentParser(description="Green IT Carbon Footprint Enrichment Pipeline")
parser.add_argument("--input", type=str, default="MYCO2 parameters dump 2026_06_10.xlsx", help="Source Excel file")
parser.add_argument("--output", type=str, default="MYCO2_parameters_Enriched.xlsx", help="Destination Excel file")
args = parser.parse_args()

SOURCE_FILE = args.input
OUTPUT_FILE = args.output
CLIMATIQ_API_KEY = os.getenv("CLIMATIQ_API_KEY", "ES1J9Y8G8S5T7A337EZ92JCD6G")

BOAVIZTA_CSV_FILE = "boavizta-data-us.csv"
PDF_FOLDER = "pcf_files"
PREFERRED_HP_REGION = "Europe"  
DEVICE_SHEET = "Devices with missing footprint"
MOBILE_SHEET = "Mobile devices with missing foo"  

KEY_COLUMN, MANUFACTURER_COLUMN, LINK_COLUMN = "DEVICE_KEY", "Manufacturer Name", "Environmental Report Link"
FOOTPRINT_COLUMN, SOURCE_COLUMN = "Total Carbon Footprint (kg CO2e)", "Data Source"
USAGE_PCT_COLUMN, SCOPE3_FOOTPRINT_COLUMN = "Usage % (Scope 2)", "Scope 3 Emissions (kg CO2e)"

SOURCE_VERIFIED_MANUAL = "Manufacturer verified (manual)"
SOURCE_VERIFIED_CATALOG = "Manufacturer verified (Boavizta catalog)"
SOURCE_ESTIMATED = "Category estimate (Boavizta)"
SOURCE_VERIFIED_LOCAL_PDF = "Manufacturer verified (local PDF)"

CATEGORY_ESTIMATES = {"laptop": 181.0, "desktop": 277.0, "tablet": 75.9, "smartphone": 84.0}

# =====================================================================
# 🌍 AUTODISCOVERY MODULE (NUBE VS LOCAL)
# =====================================================================
def get_ip_region():
    """Geolocates the server via IP. Checks manual override first."""
    # PASO 0: Leer la variable manual primero
    if os.getenv("CLOUD_REGION"): return os.getenv("CLOUD_REGION")
    
    try:
        res = requests.get("https://ipinfo.io/json", timeout=2)
        if res.status_code == 200: return res.json().get("country", "ES")
    except: pass
    return "ES"

def autodiscover_environment():
    """Intelligently detects environment, respecting manual overrides."""
    
    # PASO 0: Leer el manual override del ingeniero (Si existen, usamos estas y cortamos aquí)
    if os.getenv("CLOUD_PROVIDER") and os.getenv("CLOUD_INSTANCE_TYPE"):
        return os.getenv("CLOUD_PROVIDER"), os.getenv("CLOUD_INSTANCE_TYPE"), get_ip_region()

    # 1. Check AWS Metadata
    try:
        aws_token = requests.put("http://169.254.169.254/latest/api/token", headers={"X-aws-ec2-metadata-token-ttl-seconds": "21600"}, timeout=1)
        if aws_token.status_code == 200:
            doc = requests.get("http://169.254.169.254/latest/dynamic/instance-identity/document", headers={"X-aws-ec2-metadata-token": aws_token.text}, timeout=1).json()
            return "aws", doc.get("instanceType"), doc.get("region")
    except: pass

    # 2. Check Azure Metadata
    try:
        az_res = requests.get("http://169.254.169.254/metadata/instance?api-version=2021-02-01", headers={"Metadata": "true"}, timeout=1)
        if az_res.status_code == 200:
            az_data = az_res.json()
            return "azure", az_data["compute"]["vmSize"], az_data["compute"]["location"]
    except: pass

    # 3. Check Google Cloud (GCP) Metadata
    try:
        gcp_res = requests.get("http://metadata.google.internal/computeMetadata/v1/instance/machine-type", headers={"Metadata-Flavor": "Google"}, timeout=1)
        if gcp_res.status_code == 200:
            return "gcp", gcp_res.text.split("/")[-1], get_ip_region()
    except: pass

    # 4. Fallback to Local/Bare-Metal Architecture Scanner
    cores = multiprocessing.cpu_count()
    os_type = platform.system()
    arch = platform.machine()
    local_specs = f"Local Machine ({os_type} - {arch} - {cores} Cores)"
    return "local", local_specs, get_ip_region()

# =====================================================================
# 🔍 DATA ENRICHMENT LOGIC (PDFS & CATALOG)
# =====================================================================
def normalize_key(text): return " ".join(str(text).strip().lower().split())

def load_boavizta_catalog(csv_path):
    index = {}
    try:
        with open(csv_path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                gwp, name = row.get("gwp_total", "").strip(), row.get("name", "").strip()
                if not gwp or not name: continue
                try: index[normalize_key(f"{row.get('manufacturer', '')} {name}".strip())] = {"footprint": float(gwp), "link": row.get("sources", "").strip()}
                except: continue
    except FileNotFoundError: pass
    return index

BOAVIZTA_CATALOG = load_boavizta_catalog(BOAVIZTA_CSV_FILE)
_REGION_INDEX = {"North America": 0, "Europe": 1, "Asia Pacific": 2}[PREFERRED_HP_REGION]

PDF_PATTERNS = [
    ("HP regional", re.compile(r"total pcf\s*(\d+[.,]?\d*)\s*kg\s*co\s*2\s*e\s*(\d+[.,]?\d*)\s*kg\s*co\s*2\s*e\s*(\d+[.,]?\d*)\s*kg\s*co\s*2\s*e", re.IGNORECASE), lambda m: [float(m.group(_REGION_INDEX + 1).replace(",", "."))]),
    ("HP table", re.compile(r"co\s*2\s*e\s*\(kg\)\s*(\d+[.,]?\d*)", re.IGNORECASE), lambda m: [float(m.group(1).replace(",", "."))]),
    ("Apple new", re.compile(r"life cycle product emissions\s*\(scope\s*3\)\s*(\d+[.,]?\d*)\s*kg", re.IGNORECASE), lambda m: [float(m.group(1).replace(",", "."))]),
    ("Apple old", re.compile(r"life cycle\s*(\d+[.,]?\d*)\s*kg\s*carbon", re.IGNORECASE), lambda m: [float(m.group(1).replace(",", "."))]),
    ("Generic", re.compile(r"(\d+[.,]?\d*)\s*kg\s*co\s*2\s*e", re.IGNORECASE), lambda m: [float(m.group(1).replace(",", "."))]),
]

def extract_pdf_text(pdf_path, max_pages=12):
    reader = PdfReader(pdf_path)
    text = "".join((page.extract_text() or "") + " " for i, page in enumerate(reader.pages) if i < max_pages)
    return re.sub(r"\s+", " ", text)

def extract_pcf_values(text):
    for _, pattern, extractor in PDF_PATTERNS:
        match = pattern.search(text)
        if match: return extractor(match), _
    return None, None

def extract_usage_pct(text):
    if match := re.compile(r"(\d{1,2})\s*%\s*(?:use\b|usage\b)", re.IGNORECASE).search(text): return float(match.group(1)) / 100.0
    if match := re.compile(r"\b(?:use\b|usage\b|use phase)[^%\d]{0,40}?(\d{1,2})\s*%", re.IGNORECASE).search(text): return float(match.group(1)) / 100.0
    return 0.15

def load_pdf_folder(folder):
    index = {}
    if not os.path.isdir(folder): return index
    for filename in sorted(os.listdir(folder)):
        if not filename.lower().endswith(".pdf"): continue
        try:
            text = extract_pdf_text(os.path.join(folder, filename))
            values, _ = extract_pcf_values(text)
            if values: index[normalize_key(os.path.splitext(filename)[0].replace("_PER_", " ").strip())] = {"footprint": values[0], "usage_pct": extract_usage_pct(text), "link": f"PDF: {filename}"}
        except: continue
    return index

PDF_INDEX = load_pdf_folder(PDF_FOLDER)

def detect_category(device_key, sheet_name):
    if sheet_name == MOBILE_SHEET: return "tablet" if "ipad" in str(device_key).lower() else "smartphone"
    return "desktop" if any(k in str(device_key).lower() for k in ["elitedesk", "mac mini"]) else "laptop"

def ensure_column(ws, hmap, name):
    if name in hmap: return hmap[name]
    new_col = ws.max_column + 1
    ws.cell(1, new_col, name)
    hmap[name] = new_col
    return new_col

def process_sheet(ws):
    hmap = {c.value: c.column for c in ws[1] if c.value}
    key_col, mfg_col, lnk_col, fp_col, src_col, pct_col, s3_col = hmap[KEY_COLUMN], ensure_column(ws, hmap, MANUFACTURER_COLUMN), ensure_column(ws, hmap, LINK_COLUMN), ensure_column(ws, hmap, FOOTPRINT_COLUMN), ensure_column(ws, hmap, SOURCE_COLUMN), ensure_column(ws, hmap, USAGE_PCT_COLUMN), ensure_column(ws, hmap, SCOPE3_FOOTPRINT_COLUMN)
    
    stats = {"processed": 0}
    for row in range(2, ws.max_row + 1):
        dk = ws.cell(row, key_col).value
        if not dk or ws.cell(row, fp_col).value: continue
            
        fp_val, u_pct, src = CATEGORY_ESTIMATES[detect_category(dk, ws.title)], 0.15, SOURCE_ESTIMATED
        if normalize_key(dk) in PDF_INDEX:
            fp_val, u_pct, src = PDF_INDEX[normalize_key(dk)]["footprint"], PDF_INDEX[normalize_key(dk)]["usage_pct"], SOURCE_VERIFIED_LOCAL_PDF
        elif normalize_key(dk) in BOAVIZTA_CATALOG:
            fp_val, src = BOAVIZTA_CATALOG[normalize_key(dk)]["footprint"], SOURCE_VERIFIED_CATALOG
            
        ws.cell(row, fp_col, fp_val); ws.cell(row, pct_col, f"{int(u_pct*100)}%"); ws.cell(row, s3_col, fp_val * (1 - u_pct)); ws.cell(row, src_col, src)
        stats["processed"] += 1
    return stats

# =====================================================================
# 🚀 DUAL-API TELEMETRY & ECO-SCI METHODOLOGY
# =====================================================================
def fetch_dynamic_telemetry(provider, instance_id, region):
    dynamic_I, dynamic_M = 200.0, 16000.0  # Fallbacks
    
    try:
        res_c = requests.post("https://api.climatiq.io/data/v1/estimate", headers={"Authorization": f"Bearer {CLIMATIQ_API_KEY}", "Content-Type": "application/json"}, json={"emission_factor": {"activity_id": "electricity-supply_grid-source_supplier_mix", "region": region, "data_version": "^0"}, "parameters": {"energy": 1, "energy_unit": "kWh"}})
        if res_c.status_code == 200: dynamic_I = res_c.json().get("co2e", 0.2) * 1000
    except: pass

    if provider in ["aws", "azure", "gcp"]:
        try:
            res_b = requests.post("https://api.boavizta.org/v1/cloud/instance", json={"provider": provider, "instance_type": instance_id})
            if res_b.status_code == 200: dynamic_M = float(res_b.json().get("impacts", {}).get("gwp", {}).get("embedded", {}).get("value", 16.0)) * 1000
        except: pass
    else:
        # LOCAL MACHINE ESTIMATE: Desktop/Laptop footprint scaled to grams
        dynamic_M = CATEGORY_ESTIMATES["laptop"] * 1000

    return dynamic_I, dynamic_M

def generate_if_manifest(energy_kwh, total_transactions, execution_duration_secs):
    provider, instance_specs, region = autodiscover_environment()
    print(f"\n📡 [SMART DISCOVERY] Detected Environment: Provider={provider.upper()}, Node={instance_specs}, GeoRegion={region}")
    
    I, M_base = fetch_dynamic_telemetry(provider, instance_specs, region)
    print(f"   🟢 Grid Intensity (I): {I:.2f} g CO2e/kWh | Hardware Base (TE x RS): {M_base:.2f} g CO2e")

    # ECO-SCI ALLOCATION (Time-Share TS)
    M_allocated = M_base * (execution_duration_secs / (4 * 365 * 24 * 3600))  
    carbon_g = energy_kwh * I  
    sci_score = (carbon_g + M_allocated) / total_transactions if total_transactions > 0 else 0

    log_file = "sci_metrics_log.csv"
    headers = ["SCI Tracker ID (SU = Software Unit)", "Date", "SCI Score (g CO2e/tx)", "Total Carbon Footprint (g CO2e)", "Energy Consumed - E (kWh)", "Carbon Intensity - I (g CO2e/kWh)", "Embodied Emissions - M (g CO2e)", "Functional Unit - R (Transactions)"]
    needs_headers = not os.path.isfile(log_file) or os.path.getsize(log_file) == 0
    with open(log_file, "a", newline="", encoding="utf-8") as file:
        w = csv.writer(file)
        if needs_headers: w.writerow(headers)
        w.writerow(["SCI-SU-001", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), f"{sci_score:.12f}", f"{carbon_g:.12f}", f"{energy_kwh:.7f}", f"{I:.2f}", f"{M_allocated:.12f}", total_transactions])
    
    print(f"🏆 [ECO-SCI SCORE] Final micro-allocated footprint: {sci_score:.12f} g CO2e per tx")

def main():
    print("🌍 [GREEN SOFTWARE SERVICES] Scanning infrastructure and initializing CodeCarbon...")
    start_time = time.time()
    tracker = EmissionsTracker(project_name="myCO2_Enrichment", log_level="error")
    tracker.start()

    wb = openpyxl.load_workbook(SOURCE_FILE)
    tx = sum(process_sheet(wb[s]).get("processed", 0) for s in (DEVICE_SHEET, MOBILE_SHEET) if s in wb.sheetnames)
    wb.save(OUTPUT_FILE)
    
    emissions, duration = tracker.stop(), time.time() - start_time
    print(f"\n🌍 [REPORT] Runtime complete. Duration: {duration:.2f} seconds. Transactions enriched: {tx}")
    
    if tx > 0: generate_if_manifest(emissions, tx, duration)

if __name__ == "__main__":
    main()